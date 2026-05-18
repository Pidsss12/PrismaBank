from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse
from django.utils import timezone
from django.db.models import Count, Avg, F
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login as auth_login
from django.contrib import messages
from .models import ServiceCategory, Counter, QueueItem, BankConfig, UserProfile
from .serializers import QueueItemSerializer
from rest_framework.decorators import api_view
from rest_framework.response import Response

def landing(request):
    config = BankConfig.objects.first()
    return render(request, 'queues/landing.html', {'config': config})

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            auth_login(request, user)
            return redirect('dashboard_redirect')
    else:
        form = UserCreationForm()
    return render(request, 'registration/register.html', {'form': form})

from django.contrib.auth import logout as auth_logout

def user_logout(request):
    auth_logout(request)
    return redirect('landing')


@login_required
def dashboard_redirect(request):
    if request.user.is_superuser:
        return redirect('admin_dashboard')
    if request.user.is_staff:
        return redirect('petugas_dashboard')
    return redirect('customer_portal')

@login_required
def customer_portal(request):
    active_ticket = QueueItem.objects.filter(
        user=request.user,
        status__in=['waiting', 'calling', 'processing']
    ).order_by('-created_at').first()
    
    ahead = 0
    if active_ticket and active_ticket.status == 'waiting':
        ahead = QueueItem.objects.filter(
            category=active_ticket.category,
            status='waiting',
            created_at__lt=active_ticket.created_at
        ).count()
        
    history = QueueItem.objects.filter(user=request.user, status='completed').order_by('-finished_at')[:5]
    
    return render(request, 'queues/customer_portal.html', {
        'active_ticket': active_ticket,
        'ahead': ahead,
        'history': history
    })

def kiosk(request):
    categories = ServiceCategory.objects.all()
    # Attach waiting count to each category object for easy access in template
    for cat in categories:
        cat.waiting_count = QueueItem.objects.filter(category=cat, status='waiting').count()
    
    config = BankConfig.objects.first()
    return render(request, 'queues/kiosk.html', {
        'categories': categories,
        'config': config
    })

def take_ticket(request, category_id):
    category = get_object_or_404(ServiceCategory, id=category_id)
    today = timezone.now().date()
    last_queue = QueueItem.objects.filter(
        category=category, 
        created_at__date=today
    ).order_by('-number').first()
    
    next_number = (last_queue.number + 1) if last_queue else 1
    
    queue = QueueItem.objects.create(
        number=next_number,
        category=category,
        status='waiting',
        user=request.user if request.user.is_authenticated else None
    )
    
    return JsonResponse({
        'status': 'success',
        'number': queue.formatted_number(),
        'category': category.name,
        'waiting': QueueItem.objects.filter(category=category, status='waiting').count() - 1,
        'id': queue.id
    })

@login_required
def switch_counter(request, counter_id):
    # Unassign user from any current counter
    Counter.objects.filter(current_petugas=request.user).update(current_petugas=None, is_active=False)
    
    # Assign to new counter if not taken
    target = get_object_or_404(Counter, id=counter_id)
    if not target.current_petugas:
        target.current_petugas = request.user
        target.is_active = True
        target.save()
        return redirect('petugas_dashboard')
    else:
        messages.error(request, 'Loket tersebut sudah digunakan petugas lain.')
        return redirect('petugas_dashboard')

@login_required
def petugas_dashboard(request):
    counter = Counter.objects.filter(current_petugas=request.user).first()
    
    # Handle auto-assignment if no counter is assigned
    if not counter:
        username_lower = request.user.username.lower()
        if 'cs' in username_lower:
            counter = Counter.objects.filter(current_petugas__isnull=True, service_category__name__icontains='cs').first()
        elif 'teller' in username_lower:
            counter = Counter.objects.filter(current_petugas__isnull=True, service_category__name__icontains='teller').first()
            
        # Fallback to any free counter if no specific matching found
        if not counter:
            counter = Counter.objects.filter(current_petugas__isnull=True).first()
            
        if counter:
            counter.current_petugas = request.user
            counter.is_active = True
            counter.save()
            
    # Show all available counters that have no active petugas (allows switching role/counter)
    available_counters = Counter.objects.filter(current_petugas__isnull=True)
    if counter:
        available_counters = available_counters.exclude(id=counter.id)
            
    current_queue = QueueItem.objects.filter(petugas=request.user, status__in=['calling', 'processing']).first()
    
    active_queues = []
    completed_count = 0
    waiting_count = 0
    
    if counter:
        # Get waiting list for this counter's category - ONLY WAITING
        active_queues = QueueItem.objects.filter(
            category=counter.service_category, 
            status='waiting'
        ).order_by('created_at')
        
        completed_count = QueueItem.objects.filter(
            category=counter.service_category,
            status='completed',
            created_at__date=timezone.now().date()
        ).count()
        
        waiting_count = QueueItem.objects.filter(
            category=counter.service_category,
            status='waiting'
        ).count()

    return render(request, 'queues/petugas.html', {
        'counter': counter,
        'current_queue': current_queue,
        'active_queues': active_queues,
        'completed_count': completed_count,
        'waiting_count': waiting_count,
        'available_counters': available_counters
    })

@login_required
def call_next(request):
    # Try to find the counter where this user is active
    counter = Counter.objects.filter(current_petugas=request.user).first()
    
    # If not found, try to auto-assign one for this session if available
    if not counter:
        username_lower = request.user.username.lower()
        if 'cs' in username_lower:
            counter = Counter.objects.filter(current_petugas__isnull=True, service_category__name__icontains='cs').first()
        elif 'teller' in username_lower:
            counter = Counter.objects.filter(current_petugas__isnull=True, service_category__name__icontains='teller').first()
            
        if not counter:
            counter = Counter.objects.filter(current_petugas__isnull=True).first()
            
        if counter:
            counter.current_petugas = request.user
            counter.is_active = True
            counter.save()
    
    if not counter:
        return JsonResponse({'status': 'error', 'message': 'Maaf, Anda belum ditugaskan ke Loket manapun.'}, status=400)
    
    # Complete any current calling/processing item for this petugas
    current = QueueItem.objects.filter(petugas=request.user, status__in=['calling', 'processing']).first()
    if current:
        current.status = 'completed'
        current.finished_at = timezone.now()
        current.save()
        
    # Get next waiting in the SAME category as the counter
    next_queue = QueueItem.objects.filter(
        category=counter.service_category, 
        status='waiting'
    ).order_by('created_at').first()
    
    if next_queue:
        next_queue.status = 'calling'
        next_queue.petugas = request.user
        next_queue.counter = counter
        next_queue.called_at = timezone.now()
        next_queue.save()
        return JsonResponse({
            'status': 'success', 
            'number': next_queue.formatted_number(),
            'counter': counter.name
        })
    
    return JsonResponse({'status': 'empty', 'message': 'Tidak ada antrean lagi di kategori ini.'})

@login_required
def recall_queue(request):
    current = QueueItem.objects.filter(petugas=request.user, status__in=['calling', 'processing']).first()
    if current:
        current.called_at = timezone.now()
        current.save()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error', 'message': 'Tidak ada antrean aktif untuk dipanggil ulang.'})

@login_required
def complete_queue(request):
    current = QueueItem.objects.filter(petugas=request.user, status__in=['calling', 'processing']).first()
    if current:
        current.status = 'completed'
        current.finished_at = timezone.now()
        current.save()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'error': 'No active session'}, status=400)

def public_monitor(request):
    config = BankConfig.objects.first()
    counters = Counter.objects.all()
    history = QueueItem.objects.filter(status='completed').order_by('-finished_at')[:5]
    return render(request, 'queues/monitor.html', {
        'config': config,
        'counters': counters,
        'history': history
    })

@api_view(['GET'])
def api_monitor_data(request):
    counters = Counter.objects.all().select_related('service_category')
    # Get the most recently called item that is STILL active (calling or processing)
    latest_call = QueueItem.objects.filter(
        called_at__isnull=False,
        status__in=['calling', 'processing']
    ).order_by('-called_at').first()
    
    # Get the most recently completed item for the "Selesai" announcement
    last_completed = QueueItem.objects.filter(status='completed').order_by('-finished_at').first()

    counter_data = []
    for c in counters:
        current = QueueItem.objects.filter(counter=c, status__in=['calling', 'processing']).order_by('-called_at').first()
        waiting = QueueItem.objects.filter(
            category=c.service_category,
            status='waiting'
        ).order_by('created_at')[:5]
        
        counter_data.append({
            'counter_id': c.id,
            'current_number': current.formatted_number() if current else '--',
            'is_active': current is not None,
            'waiting_list': [item.formatted_number() for item in waiting]
        })

    latest_call_data = {
        'id': latest_call.id,
        'formatted_number': latest_call.formatted_number(),
        'counter_name': latest_call.counter.name,
        'counter_id': latest_call.counter.id,
        'called_at': latest_call.called_at.isoformat() if latest_call.called_at else None
    } if latest_call else None
    
    return Response({
        'counters': counter_data,
        'latest_call': latest_call_data,
        'last_completed_id': last_completed.id if last_completed else None
    })

@user_passes_test(lambda u: u.is_superuser)
def export_queues_csv(request):
    from django.http import HttpResponse
    from django.utils import timezone
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from collections import defaultdict

    # Retrieve all queues sorted by created_at (ascending) so dates flow chronologically
    queues = QueueItem.objects.all().select_related('category', 'petugas').order_by('created_at')
    
    # Group queues by local date
    queues_by_date = defaultdict(list)
    for q in queues:
        local_date = timezone.localtime(q.created_at).date()
        queues_by_date[local_date].append(q)

    # Create Workbook
    wb = Workbook()
    
    # Style definitions in CFI (Corporate Finance Institute) Premium Theme
    fill_cfi_navy = PatternFill(start_color='0C2340', end_color='0C2340', fill_type='solid') # CFI Deep Navy
    fill_accent_gold = PatternFill(start_color='FF9E1B', end_color='FF9E1B', fill_type='solid') # CFI Accent Gold
    
    font_brand = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    font_cfi_header = Font(name='Segoe UI', size=10, bold=True, color='0C2340')
    font_cfi_title = Font(name='Segoe UI', size=14, bold=True, color='0C2340')
    font_cfi_subtitle = Font(name='Segoe UI', size=9, italic=True, color='475569')
    
    font_header = Font(name='Segoe UI', size=9, bold=True, color='0C2340')
    font_data = Font(name='Segoe UI', size=9, color='334155')
    
    fill_zebra = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid') # Very subtle slate grey
    fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Status badges (matching CFI's elegant pastel highlight tones)
    fill_completed = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid') # soft green
    font_completed = Font(name='Segoe UI', size=8, bold=True, color='1B5E20')

    fill_waiting = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')   # soft amber
    font_waiting = Font(name='Segoe UI', size=8, bold=True, color='B78103')

    fill_calling = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')   # soft blue
    font_calling = Font(name='Segoe UI', size=8, bold=True, color='0D47A1')

    fill_missed = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')     # soft red
    font_missed = Font(name='Segoe UI', size=8, bold=True, color='C62828')
    
    # CFI border styles (No vertical lines, medium navy highlights for headers/footers)
    border_thin_bottom = Side(border_style='thin', color='E2E8F0')
    border_medium_navy = Side(border_style='medium', color='0C2340')
    border_double_navy = Side(border_style='double', color='0C2340')
    
    border_header = Border(top=border_thin_bottom, bottom=border_medium_navy, left=None, right=None)
    border_data = Border(bottom=border_thin_bottom, left=None, right=None)
    border_last_row = Border(bottom=border_double_navy, left=None, right=None)
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # If there are no queues, create a default empty sheet
    if not queues_by_date:
        ws = wb.active
        ws.title = "No Data"
        ws.append(['Tidak ada data antrean'])
    else:
        # Sort dates ascending so early dates appear as sheet 1, and so on
        sorted_dates = sorted(queues_by_date.keys())
        first_sheet = True

        for date_key in sorted_dates:
            sheet_name = date_key.strftime('%d-%m-%Y')
            
            if first_sheet:
                ws = wb.active
                ws.title = sheet_name
                first_sheet = False
            else:
                ws = wb.create_sheet(title=sheet_name)
                
            # Set gridlines visible (standard in CFI except for merged titles)
            ws.views.sheetView[0].showGridLines = True
            
            # --- 1. WRITE SOLID NAVY BRAND BANNER ---
            ws.row_dimensions[2].height = 30
            ws.merge_cells('A2:G2')
            
            # Apply navy fill across the merged block
            for col in range(1, 8):
                cell = ws.cell(row=2, column=col)
                cell.fill = fill_cfi_navy
                
            # Put title text in A2
            title_cell = ws.cell(row=2, column=1, value="PRISMA BANKING SYSTEM  —  DAILY PERFORMANCE JOURNAL")
            title_cell.font = font_brand
            title_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            
            # Subtitle information below the banner
            ws.row_dimensions[4].height = 20
            info_cell = ws.cell(row=4, column=1, value=f"Tanggal Operasional: {sheet_name}   |   Kantor Cabang: Utama Jakarta   |   Klasifikasi: Internal Audit")
            info_cell.font = font_cfi_subtitle
            
            # Calculate metrics for the day
            day_queues = queues_by_date[date_key]
            total_day = len(day_queues)
            completed_day = len([q for q in day_queues if q.status == 'completed'])
            missed_day = len([q for q in day_queues if q.status == 'missed'])
            
            # Calculate average wait time for the day
            wait_durations = []
            for q in day_queues:
                if q.called_at and q.created_at:
                    wait_durations.append((q.called_at - q.created_at).total_seconds())
            if wait_durations:
                avg_seconds = sum(wait_durations) / len(wait_durations)
                avg_wait_str = f"{int(avg_seconds // 60)}m {int(avg_seconds % 60)}s"
            else:
                avg_wait_str = "0m 0s"
                
            # --- 2. WRITE CFI METRICS SUMMARY SIDE-BLOCK (COLUMNS I TO K) ---
            # Header Card
            ws.row_dimensions[6].height = 22
            ws.merge_cells('I6:K6')
            for col in range(9, 12):
                cell = ws.cell(row=6, column=col)
                cell.fill = fill_cfi_navy
            card_header = ws.cell(row=6, column=9, value="KEY PERFORMANCE INDICATORS")
            card_header.font = Font(name='Segoe UI', size=8, bold=True, color='FFFFFF')
            card_header.alignment = center_align
            
            metrics = [
                ("Total Tiket Antrean", total_day, 'FFF3E0', 'C2410C'),       # Soft Gold
                ("Tiket Terselesaikan", completed_day, 'E8F5E9', '1B5E20'),   # Soft Green
                ("Tiket Terlewat (Missed)", missed_day, 'FFEBEE', 'C62828'), # Soft Red
                ("Rata-rata Waktu Tunggu", avg_wait_str, 'E3F2FD', '0D47A1')  # Soft Blue
            ]
            
            for idx, (label, val, bg_color, text_color) in enumerate(metrics, 7):
                ws.row_dimensions[idx].height = 20
                ws.merge_cells(start_row=idx, start_column=9, end_row=idx, end_column=10)
                
                # Style label cell
                lbl_cell = ws.cell(row=idx, column=9, value=label)
                lbl_cell.font = Font(name='Segoe UI', size=8, bold=True, color='475569')
                lbl_cell.alignment = left_align
                lbl_cell.border = Border(bottom=border_thin_bottom)
                
                # Apply bottom border to column 10 (which is merged with 9)
                ws.cell(row=idx, column=10).border = Border(bottom=border_thin_bottom)
                
                # Style value cell
                val_cell = ws.cell(row=idx, column=11, value=val)
                val_cell.font = Font(name='Segoe UI', size=8, bold=True, color=text_color)
                val_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                val_cell.alignment = center_align
                val_cell.border = Border(bottom=border_thin_bottom)
            
            # --- 3. WRITE SEQUENTIAL DATA TABLES (COLUMNS A TO G) ---
            categories = ServiceCategory.objects.all()
            current_row = 6
            
            for cat in categories:
                cat_queues = [q for q in day_queues if q.category_id == cat.id]
                
                if not cat_queues:
                    continue
                
                # Write Category Section Header
                ws.cell(row=current_row, column=1, value=f"KATEGORI LAYANAN: {cat.name.upper()}").font = Font(name='Segoe UI', size=10, bold=True, color='0C2340')
                ws.row_dimensions[current_row].height = 24
                # Subtle bottom line accent for category name
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).border = Border(bottom=Side(style='thin', color='CBD5E1'))
                current_row += 1
                
                # Write headers
                headers = ['Nomor Antrean', 'Kategori Layanan', 'Status Antrean', 'Jam Dibuat', 'Jam Dipanggil', 'Jam Selesai', 'Petugas Loket']
                ws.row_dimensions[current_row].height = 26
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=header)
                    cell.font = font_header
                    cell.border = border_header
                    cell.alignment = center_align
                
                current_row += 1
                
                # Write data rows
                start_data_row = current_row
                num_rows = len(cat_queues)
                last_row_idx = start_data_row + num_rows - 1
                
                for idx, q in enumerate(cat_queues):
                    ws.row_dimensions[current_row].height = 20
                    
                    created = timezone.localtime(q.created_at).strftime('%H:%M:%S')
                    called = timezone.localtime(q.called_at).strftime('%H:%M:%S') if q.called_at else '-'
                    finished = timezone.localtime(q.finished_at).strftime('%H:%M:%S') if q.finished_at else '-'
                    
                    row_fill = fill_zebra if idx % 2 == 1 else fill_white
                    
                    row_values = [
                        q.formatted_number(),
                        q.category.name,
                        q.status.upper(),
                        created,
                        called,
                        finished,
                        q.petugas.username if q.petugas else '-'
                    ]
                    
                    current_border = border_last_row if current_row == last_row_idx else border_data
                    
                    for col_num, val in enumerate(row_values, 1):
                        cell = ws.cell(row=current_row, column=col_num, value=val)
                        cell.font = font_data
                        cell.fill = row_fill
                        cell.border = current_border
                        cell.alignment = center_align if col_num != 2 and col_num != 7 else left_align
                        
                        # Apply elegant CFI-style Status Badge
                        if col_num == 3:
                            status_str = val.lower()
                            if 'completed' in status_str:
                                cell.fill = fill_completed
                                cell.font = font_completed
                            elif 'waiting' in status_str:
                                cell.fill = fill_waiting
                                cell.font = font_waiting
                            elif 'calling' in status_str or 'progress' in status_str:
                                cell.fill = fill_calling
                                cell.font = font_calling
                            elif 'missed' in status_str:
                                cell.fill = fill_missed
                                cell.font = font_missed
                                
                    current_row += 1
                
                # Spacer row before the next category table
                current_row += 2
            
            # Dynamic auto-column-width adjustment (excluding title rows)
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                
                # Check if it's the metrics side-panel columns (I, J, K) to set fixed elegant widths
                if col_letter in ['I', 'J', 'K']:
                    if col_letter == 'I':
                        ws.column_dimensions[col_letter].width = 18
                    elif col_letter == 'J':
                        ws.column_dimensions[col_letter].width = 10
                    elif col_letter == 'K':
                        ws.column_dimensions[col_letter].width = 14
                    continue
                    
                for cell in col:
                    # Ignore banner row 2, 3, 4 for width calculations
                    if cell.row in [2, 3, 4]:
                        continue
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 6, 18)

    # Set up HTTP response as Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Prisma_Bank_Queue_Report.xlsx"'
    wb.save(response)
    
    return response

@user_passes_test(lambda u: u.is_superuser)
def admin_dashboard(request):
    from django.db.models import Count, Avg, F, ExpressionWrapper, fields
    
    stats = {
        'total_queues': QueueItem.objects.count(),
        'completed': QueueItem.objects.filter(status='completed').count(),
        'missed': QueueItem.objects.filter(status='missed').count(),
        'by_category': QueueItem.objects.values('category__name').annotate(count=Count('id')),
    }
    
    # Improved Wait Time calculation
    avg_wait = QueueItem.objects.filter(status='completed', called_at__isnull=False).annotate(
        wait_duration=ExpressionWrapper(F('called_at') - F('created_at'), output_field=fields.DurationField())
    ).aggregate(avg_wait=Avg('wait_duration'))
    
    # Format wait time for display
    if avg_wait['avg_wait']:
        seconds = int(avg_wait['avg_wait'].total_seconds())
        minutes = seconds // 60
        stats['formatted_wait'] = f"{minutes}m {seconds % 60}s"
    else:
        stats['formatted_wait'] = "0m 0s"
        
    recent_activity = QueueItem.objects.all().order_by('-created_at')[:10]
    
    return render(request, 'queues/admin.html', {
        'stats': stats,
        'recent_activity': recent_activity
    })

def queue_status(request, item_id):
    from django.shortcuts import get_object_or_404
    item = get_object_or_404(QueueItem, id=item_id)
    # Hitung berapa orang di depan (yang statusnya 'waiting' dan ID lebih kecil)
    ahead = QueueItem.objects.filter(
        category=item.category,
        status='waiting',
        id__lt=item.id
    ).count()
    
    return render(request, 'queues/status.html', {
        'item': item,
        'ahead': ahead
    })


@login_required
def update_profile(request):
    from django.contrib.auth.models import User
    if request.method == 'POST':
        user = request.user
        username = request.POST.get('username')
        first_name = request.POST.get('first_name')
        
        # Validation
        if username:
            # Check if username already exists for another user
            if User.objects.filter(username=username).exclude(id=user.id).exists():
                return JsonResponse({'status': 'error', 'message': 'Username sudah digunakan oleh orang lain.'}, status=400)
            user.username = username
            
        if first_name is not None:
            user.first_name = first_name
            
        user.save()
        
        profile, _ = UserProfile.objects.get_or_create(user=user)
        if 'avatar' in request.FILES:
            # Delete old avatar if exists
            if profile.avatar:
                profile.avatar.delete(save=False)
            profile.avatar = request.FILES['avatar']
            profile.save()
            
        return JsonResponse({
            'status': 'success', 
            'message': 'Profil berhasil diperbarui!',
            'username': user.username,
            'first_name': user.first_name,
            'avatar_url': profile.avatar.url if profile.avatar else None
        })
        
    return JsonResponse({'status': 'error', 'message': 'Metode tidak diperbolehkan.'}, status=405)
